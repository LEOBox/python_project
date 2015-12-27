#-*- coding: UTF-8 -*-
from openpyxl import load_workbook
import json
from lxml import etree

def readData(filename):
	wb = load_workbook(filename)
	ws = wb.get_sheet_by_name('Sheet1')
	row = ws.max_row
	col = ws.max_column
	data = []
	for i in range(row):
		temp = []
		for j in range(col):
			temp.append(ws.rows[i][j].value)
		data.append(temp)
	return json.dumps(data,encoding = 'utf-8')

def writeXML(jsonData):
	root = etree.Element('root')
	numbers = etree.SubElement(root,'numbers')
	comment = etree.Comment(u"""数字""")
	numbers.append(comment)
	numbers.text = jsonData
	numbersXML = etree.ElementTree(root)
	numbersXML.write('numberss.xml',pretty_print = True,xml_declaration = True, encoding = 'utf-8')

if __name__ == '__main__':
	writeXML(readData('numbers.xlsx'))