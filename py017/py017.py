#-*- coding: UTF-8 -*-
from openpyxl import load_workbook
import json
from lxml import etree

def readData(filename):
	wb = load_workbook(filename)
	ws = wb.get_sheet_by_name('Sheet1')
	row = ws.max_row
	col = ws.max_column
	data = {}
	for i in range(0,row):
		for j in range(1,col):
			data.setdefault(ws.rows[i][0].value,[]).append(ws.rows[i][j].value)
	dataJson = json.dumps(data,encoding = 'utf-8')
	return dataJson

def writeXML(dataJson):
	root = etree.Element('root')
	students = etree.SubElement(root,'student')
	comment = etree.Comment(u"""StudentInfo"id":[name，math,art,english]""")
	students.append(comment)
	students.text = dataJson
	studentXML = etree.ElementTree(root)
	studentXML.write('student.xml',pretty_print = True,xml_declaration = True, encoding = 'utf-8')
	

if __name__ == '__main__':
	writeXML(readData('student.xlsx'))
#	read_exl('student.xlsx')
