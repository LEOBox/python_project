from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

import json

def openTxtFile(filePath):
	f = open(filePath)
	fileText = f.read().decode('utf-8')
	jsonType = json.loads(fileText)
	return jsonType
#	print jsonType

def writeExcel(jsonType):
	wb = Workbook()
	ws = wb.create_sheet(0,0)
	row = 1
	col = 1
	for number in sorted(jsonType.items(),key = lambda d:d[0]):
		for info in number:
			colLetter = get_column_letter(col)
			ws.cell('%s%s'%(colLetter,row)).value = info
			col += 1
		row += 1
		col = 1
	wb.save('./cities.xlsx')
	print 'OK'


if __name__ == '__main__':
	 text = openTxtFile('./cities.txt')
	 writeExcel(text)

