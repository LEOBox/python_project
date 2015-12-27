#-*- coding: UTF-8 -*-
from openpyxl import load_workbook
import json
from lxml import etree

def readData(filename):
	wb = load_workbook(filename)
	ws = wb.get_sheet_by_name('Sheet1')
	row = ws.max_row
	col = ws.max_column
	data = {}
	for i in range(row):
		for j in range(1,col):
			data.setdefault(ws.rows[i][0].value,[]).append(ws.rows[i][j].value)
	return json.dumps(data,encoding = 'utf-8')

def writeXML(jsonData):
	root = etree.Element('root')
	cities = etree.SubElement(root,'cities')
	comment = etree.Comment(u"""城市""")
	cities.append(comment)
	cities.text = jsonData
	citiesXML = etree.ElementTree(root)
	citiesXML.write('cities.xml',pretty_print = True,xml_declaration = True, encoding = 'utf-8')

if __name__ == '__main__':
	writeXML(readData('cities.xlsx'))